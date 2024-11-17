import requests
import os
from pathManager import PathManager
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

class UpdateStats:
    def __init__(self, dossierStats,fichier_sortie, fichier_joueurs):
        self.dossier = dossierStats
        self.fichier_sortie = fichier_sortie
        self.fichier_joueurs = fichier_joueurs

    def lire_fichier_joueurs(self):
        joueurs_urls = []
        with open(self.fichier_joueurs, 'r') as f:
            lines = f.readlines()
            for line in lines:
                parts = line.strip().split(';')
                if len(parts) == 2:
                    nom, url = parts
                    joueurs_urls.append((nom.strip(), url.strip()))
                else:
                    print(f"Ligne mal formatée (ignorée) : {line.strip()}")
        return joueurs_urls

    def grab_and_write(self, url_joueur, output_file, joueur):
        response = requests.get(url_joueur)
        if response.status_code == 200:
            html_content = response.text
            soup = BeautifulSoup(html_content, 'html.parser')
            section = soup.find('section', class_='table-section', id='regul')
            headers = []
            rows = []

            if section:
                table = section.find('table')
                if table:
                    header_row = table.find('thead').find_all('th')
                    headers = [header.text.strip() for header in header_row]
                    body_rows = table.find('tbody').find_all('tr')
                    for row in body_rows:
                        columns = row.find_all('td')
                        data = [column.text.strip() for column in columns]
                        rows.append(data)
                    df = pd.DataFrame(rows, columns=headers)
                    df = df.iloc[:, list(range(10)) + [11]]
                    for col in df.columns[3:7]:
                        df[col] = pd.to_numeric(df[col].str.replace('\n', ''), errors='coerce')
                    col = df.columns[7]
                    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Stats')
                        workbook = writer.book
                        sheet = writer.sheets['Stats']
                        sheet.insert_rows(1)
                        sheet.cell(row=1, column=1, value=joueur)
                        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
                        sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
                    print(f"Le fichier Excel a été sauvegardé sous : {output_file}")
                else:
                    print("Table with the specified class not found.")
            else:
                print("Section with the specified ID not found.")
        else:
            print(f"Échec de la requête, code de statut : {response.status_code}")

    def check_and_write(self):
        joueurs_urls = self.lire_fichier_joueurs()
        for joueur, url in joueurs_urls:
            output_excel_file = os.path.join(self.dossier, f'{joueur}.xlsx')
            self.grab_and_write(url, output_excel_file, joueur)

    def lire_fichiers_du_dossier(self):
        fichiers_excel = [os.path.join(self.dossier, f) for f in os.listdir(self.dossier) if f.endswith('.xlsx')]
        return fichiers_excel

    def regrouper_statistiques(self):
        with pd.ExcelWriter(self.fichier_sortie, engine='openpyxl') as writer:
            fichiers_excel = self.lire_fichiers_du_dossier()
            for fichier in fichiers_excel:
                df = pd.read_excel(fichier, engine='openpyxl')
                nom_joueur = os.path.basename(fichier).replace('.xlsx', '')
                df.to_excel(writer, sheet_name=nom_joueur, index=False)
            print(f"Le fichier Excel combiné a été sauvegardé sous : {self.fichier_sortie}")

# Utilisation de la classe UpdateStats
P = PathManager()
dossier_stats = P.getDossierStats()
fichier_sortie = P.getFichierDossierData('Statistiques_Joueurs.xlsx')
fichier_joueurs = P.getFichierJoueurs()

update_stats = UpdateStats(dossier_stats, fichier_sortie, fichier_joueurs)
update_stats.check_and_write()
update_stats.regrouper_statistiques()